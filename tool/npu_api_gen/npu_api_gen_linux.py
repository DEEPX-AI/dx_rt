#import xlrd
import os
import sys
import numpy as np
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from copy import copy
from mako.template import Template
from mako.lookup import TemplateLookup
from requests import post
from generate_template import *
import operator
import re

NDF = 'NotDefined'
EXCEL_POSTFIX = '.xlsx'
SPLIT_EXCEL = 'split.xlsx'
npu_mem = ['Feature Mem', 'Weight0 Mem', 'Weight1 Mem', 'BN Mem', 'Bias Mem']
npu_memmap = {}
sys_memmap = {}
info_str = {}
info_num = {}

def get_mem_type(type):
    return {
        0 : "N_MEM_NONE",
        1 : "N_MEM_R",
        2 : "N_MEM_W",
        3 : "N_MEM_RW",
    }.get(type, "N_MEM_NONE")

#RegBits -> Reg -> Module
class RegBits:
    def __init__(self):
        self.category = NDF
        self.name = NDF
        self.addr = -1
        self.field = NDF
        self.onoff = NDF
        self.start_bit = -1
        self.end_bit = -1
        self.width = -1
        self.bits = NDF
        self.bits_pos = NDF
        self.type = NDF
        self.base = NDF
        self.reset = -1
        self.desc = NDF
        self.perm = NDF # added manually
        self.read = NDF # added manually
        self.write = NDF # added manually
        self.enable = NDF # added manually
    def __init__(self, list):
        self.category = list[1].replace('\n','').strip().upper()
        self.name = list[2].strip().upper()
        self.addr = int(list[3])
        self.field = list[4].strip().upper()
        self.onoff = list[5].strip().upper()
        self.start_bit = int(list[6])
        self.width = int(list[7])
        self.end_bit = self.start_bit - self.width + 1
        self.mask_bits_str = '0b'+'1'*self.width
        self.mask_bits = int(self.mask_bits_str,2)
        self.mask = self.mask_bits<<self.end_bit
        self.bits = list[8].strip().upper()
        self.bits_pos = list[9].strip().upper()
        self.type = list[10].strip().upper()
        self.base = list[11].strip().upper()
        if self.base=='HEX':
            self.reset = int('0x'+list[12].strip().upper(), 16)
        elif self.base=='DEC':
            self.reset = int(list[12].strip().upper()) # TODO
        self.desc = list[13].strip().upper().replace('\n', ' ')
        if self.type=='RW':
            self.perm = 'RW'
            self.read = 'true'
            self.write = 'true'
        elif self.type=='R':
            self.perm = 'RO'
            self.read = 'true'
            self.write = 'false'
        elif self.type=='W':
            self.perm = 'WO'
            self.read = 'false'
            self.write = 'true'
        if self.onoff == 'ON':
            self.enable = 'true'
        else:
            self.enable = 'false'
    def __str__ (self):
        return '    ' + str(
            [self.category , self.name , self.addr , self.field , self.onoff , 
            self.start_bit , self.width , self.bits , self.bits_pos , self.type , 
            self.base , self.reset , self.desc]
        ) + '\n'

class Reg:
    def __init__(self):
        self.module_name = NDF
        self.category = NDF
        self.name = NDF
        self.addr = -1
        self.regbits = []
        self.reserved_cnt = 0
        self.has_off_bit = 'false'
    def __str__(self):
        msg = f'Module {self.module_name}      Register {self.name}      Base {self.addr}({hex(self.addr)})\n'
        for regbits in self.regbits:
            msg += str(regbits)
        return msg

class Module:
    def __init__(self):
        self.baseaddr = -1
        self.name = NDF
        self.regs = []
        self.id = 0
        self.postfix = ''
    def __str__(self):
        msg = f'Module {self.name}      Base {self.baseaddr}({hex(self.baseaddr)})\n'
        msg += 'Registers ' + str([reg.name for reg in self.regs]) + '\n'
        for reg in self.regs:
            msg += str(reg)
        return msg

class Memory:
    def __init__(self, list):
        self.type   = list[1]
        self.name   = list[2]
        self.width  = int(list[3])
        self.depth  = int(list[4])
        self.size   = int(list[5])
        self.start  = int(list[6],16)
        self.end    = int(list[7],16)
        self.user_dma_rd = int(list[8])
        self.user_dma_wr = int(list[9])
        self.user_dma_cp = int(list[10])
        self.user_dma_bc = int(list[11])
        if len(list)==13:
            self.npu_ppu_mem = int(list[12])
        self.mem_type   = get_mem_type(self.user_dma_rd + (self.user_dma_wr<<1))
        self.cpu_mem    = 0
        self.group_id   = -1
        self.peth       = -1
        self.id         = 0
        self.postfix    = ''
        # print(f'Memory:{self.name}\t Size:{self.size}\t Start:{hex(self.start)}\t End:{hex(self.end)}\n')
        # assert len(list)==13, f"Please Check Excel( size : {len(list)} )"
        assert (((self.end+1)-self.start)*(self.width) == self.width*self.depth) and (self.size == self.width*self.depth/8/1024),\
                f"Please Check Address ( name:{self.name}, end:{hex(self.end)}, start:{hex(self.start)}, size:{self.size})\
                  \n    calculated size1 : {hex(((self.end+1)-self.start)*64)}, calculatated size2 : {hex(self.size*1024)}"

def pr_debug(*x):
    if verbose is True:
        print(*x)

def split_merged_cells(input, output):
    wb = load_workbook(filename = input, data_only = True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        mcr_coord_list = [mcr.coord for mcr in sheet.merged_cells.ranges]

        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            top_left_cell_alignment = sheet.cell(row=min_row, column=min_col).alignment
            top_left_cell_format = sheet.cell(row=min_row, column=min_col).number_format
            sheet.unmerge_cells(mcr)
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
                    cell.alignment = copy(top_left_cell_alignment)
                    cell.number_format = copy(top_left_cell_format)
    wb.save(output)

def get_string_except_lastnum(string):
    return re.sub(r'\d+', '', string)

def generate_module_mem(filename, module_name, postfix='', id='', npu_variant=''):
    memorys = []
    rawdata = pd.read_excel(filename,
            sheet_name='_'.join([module_name, postfix]) if postfix != '' else module_name,
            header=4,
            keep_default_na=False,
            dtype = {
                'Name'   : str,
                'Width'     : str,
                'Depth'     : str,
                'Size(KB)'  : str,
                'Start'     : str,
                'End'       : str
                }).values.tolist()
    for data in rawdata:
        mem = Memory(data)
        mem.postfix = postfix
        mem.id = id
        mem.name = 'MEM' + mem.id + '_' + mem.name
        memorys.append(mem)
    ## sort by memory name to set group id, peth
    # memorys.sort(key=operator.attrgetter('name'))
    print(f"memorys[0].name: {str(memorys[0].name)} type:{str(memorys[0].type)}")
    groud_id = 0
    mem_name_prev = ""
    for mem in memorys:
        if "DX_L2" in npu_variant:
            if 'Variable Memory' in mem.type:
                mem.group_id = groud_id
                groud_id += 1
            else: # 'RF Memory'
                mem_name = get_string_except_lastnum(mem.name)
                if mem_name != mem_name_prev and mem_name_prev != "":
                    groud_id += 1
                mem.group_id = groud_id
                mem_name_prev = mem_name
        else:
            mem.group_id = groud_id
            if 'CPU Memory' in mem.type:
                mem.cpu_mem = 1
            if mem.user_dma_cp == 0: #DX_M1A
                continue
            if (int(mem.width/8) >= 64):
                groud_id += 1
        # mem.peth = str(mem.name).split('_')[0].replace('PETH','')
        print(f"{mem.name} : {mem.group_id} {mem.peth}")
    # memorys.sort(key=operator.attrgetter('start'))
    for mem in memorys:
        mem.name = npu_variant+'_'+mem.name

    return memorys

def generate_module(filename, module_name, postfix='', id='', npu_variant=''):
    reg_def[module_name] = {}
    rawdata = pd.read_excel(filename,
                sheet_name='_'.join([module_name, postfix]) if postfix != '' else module_name,
                header=3,
                keep_default_na=False,
                dtype={'Category': str, 'NAME': str,
                    'Address': np.int32, 'Field': str,
                    'On/Off': str, 'Start Bit': np.int32,
                    'Width': np.int32, 'Bits': str,
                    'Bits_pos': str, 'Type': str,
                    'Base': str, 'Reset': str,
                    'Description': str
                    }).values.tolist()
    for data in rawdata:
        regbits = RegBits(data)
        if reg_def[module_name].get(regbits.name):
            reg_def[module_name][regbits.name].append(regbits)
        else:
            if regbits.category != "HWCONLY":
                pr_debug(f'create [{module_name}][{regbits.category}][{regbits.name}]')
                reg_def[module_name][regbits.name]=[regbits]

    module = Module()
    module.id = id
    module.postfix = postfix
    module.baseaddr = int(pd.read_excel(filename,
                                    sheet_name='_'.join([module_name, postfix]) if postfix != '' else module_name
                                    ).loc[0][2], 16)
    module.name = module_name if module_name not in 'NPU' else 'NPU'+module.id 
    org_modulename = module.name
    # module.name = npu_variant + '_' + module.name
    module.regs = []
    for reg_name in reg_def[module_name].keys():
        reg = Reg()
        reg.module_name = module_name
        bit_sum = 0
        for regbits in reg_def[module_name][reg_name]:
            if regbits.enable == 'false':
                reg.has_off_bit = 'true'
        for regbits in reg_def[module_name][reg_name]:
            regbits.category = org_modulename
            if reg.category==NDF:
                reg.category = regbits.category
            if reg.name==NDF:
                reg.name = regbits.name
            if reg.addr==-1:
                reg.addr = regbits.addr
            if regbits.field=='RESERVED':
                regbits.field += str(reg.reserved_cnt)
                reg.reserved_cnt += 1
            reg.regbits.append(regbits)
            bit_sum += regbits.width
        if bit_sum!=32:
            print(f"Error: bit sum of {module_name}-{reg.name} is not 32bit")
            exit(-1)
        reg.regbits.reverse()
        # pr_debug(module_name, reg.name)
        # for regbits in reg.regbits:
        #     pr_debug("    ", regbits)
        # print(reg)
        # reg.name = npu_variant + '_' + reg_name
        module.regs.append(reg)
    return module

verbose = True
reg_def = {}

if __name__ == '__main__':
    input_excel = sys.argv[1]
    output_dir = input_excel.split(EXCEL_POSTFIX)[0]
    npu_variant = output_dir
    # output_dir = 'generated'
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir)
    os.makedirs( f'{output_dir}/linux' )
    print(f'Extract {input_excel}......')
    split_merged_cells(input_excel, SPLIT_EXCEL)
    sheet_names = load_workbook(filename = SPLIT_EXCEL).sheetnames
    postfixes = [y for y in [x.split('_')[1] if "Memory" in x and x!="Memory" else "" for x in sheet_names] if y!='']
    if len(postfixes)==0:
        postfixes = [""]
    print(postfixes)
    mems, npus = [], []
    for idx in range(len(postfixes)):
        print(idx)
        postfix = postfixes[idx]
        mems.append(generate_module_mem(SPLIT_EXCEL, 'Memory', postfix, str(idx), npu_variant))
        npus.append(generate_module(SPLIT_EXCEL, 'NPU', postfix, str(idx), npu_variant=npu_variant))
        # memory = generate_module_mem(SPLIT_EXCEL, '_'.join(['Memory',mac_num]) if mac_num != '' else 'Memory')
    system = generate_module(SPLIT_EXCEL, 'SYSTEM', npu_variant=npu_variant)
    debug = generate_module(SPLIT_EXCEL, 'DEBUG', npu_variant=npu_variant)
    dma = generate_module(SPLIT_EXCEL, 'DMA', npu_variant=npu_variant)
    # npu = generate_module(SPLIT_EXCEL, 'NPU', mac_num)

    pr_debug(mems)
    pr_debug(system)
    pr_debug(debug)
    pr_debug(dma)
    pr_debug(npus)

    generate_template_module(input_excel, 'template/linux/npu_reg.mako', f'{output_dir}/linux/npu_reg_sys_{npu_variant}.h', system, npu_variant)
    generate_template_module(input_excel, 'template/linux/npu_reg.mako', f'{output_dir}/linux/npu_reg_dma_{npu_variant}.h', dma, npu_variant)
    generate_template_module(input_excel, 'template/linux/npu_reg.mako', f'{output_dir}/linux/npu_reg_debug_{npu_variant}.h', debug, npu_variant)

    # # if os.path.isfile(SPLIT_EXCEL):
    # #     os.remove(SPLIT_EXCEL)

